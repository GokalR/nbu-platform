"""Phase 2A: importer for the business directory.

Loads three Excel files from `chatbot/business/` into the chatbot Postgres:
- `OKED.xlsx` → `business_companies` (~167k rows)
- `TN_VED.xlsx` "Ойоғи" sheet → `business_imports` (~427k rows)
- `TN_VED.xlsx` "Мижозлар бўйича" sheet → `business_import_summaries` (~3k)
- `TN_VED_description.xls` → `tnved_categories` (~28k rows)

A single `import_runs` row covers the whole business import. To keep it out of
the way of the CERR `WHERE status='completed'` latest-run filter, this run is
marked `status='completed_business'`. The business views in `db/views.py` look
for that status when resolving "latest business snapshot".

Performance: streaming reads (openpyxl `read_only=True`) + Core bulk inserts in
batches of 2000 rows. The whole import takes a few minutes on a Railway
managed Postgres over the public proxy.

Geography: company / importer addresses are matched against the existing
CERR `regions` / `districts` / `mahallas` tables by Cyrillic-normalised name.
Unmatched rows still get inserted, just with NULL geo FK columns.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator

from sqlalchemy import Engine, select, text
from sqlalchemy.orm import Session

from cerr_chatbot.config import Settings, get_settings
from cerr_chatbot.db import (
    BusinessCompany,
    BusinessImport,
    BusinessImportSummary,
    District,
    ImportRun,
    Mahalla,
    Region,
    TnvedCategory,
)

# Bulk insert batch size. 2000 rows fits well within Railway's connection
# packet limits and keeps memory bounded.
_BATCH_SIZE = 2000

# Suffixes stripped before matching geographic names. Order matters: longer
# suffixes first so " вилояти" doesn't shadow " лояти" inside " вилояти".
_GEO_SUFFIXES = (
    " вилояти",
    " шаҳри",
    " шахри",
    " шаҳар",
    " шахар",
    " тумани",
    " туман",
    " республикаси",
)

# Map Uzbek-specific Cyrillic glyphs to base Russian Cyrillic so common
# misspellings still resolve.
_GLYPH_FOLD = str.maketrans(
    {"қ": "к", "ҳ": "х", "ў": "у", "ғ": "г", "ё": "е", "Қ": "к", "Ҳ": "х", "Ў": "у", "Ғ": "г"}
)


def _normalize_geo_name(raw: Any) -> str:
    """Lower-case, strip common geo suffixes, fold Uzbek-specific glyphs.

    Used on both sides of the name-match so source data spelled "ҚАШҚАДАРЁ
    ВИЛОЯТИ" resolves to the CERR snapshot's "Қашқадарё вилояти" (and
    similar variants spelled with plain к / х / у / г).
    """
    if not isinstance(raw, str):
        return ""
    s = raw.strip().lower()
    if not s:
        return ""
    s = s.translate(_GLYPH_FOLD)
    # Strip suffixes (one pass is fine; names like "Тошкент шаҳри тумани" do
    # not occur in the source).
    for suf in _GEO_SUFFIXES:
        if s.endswith(suf):
            s = s[: -len(suf)]
            break
    s = re.sub(r"\s+", " ", s).strip(" .-")
    return s


@dataclass
class BusinessImportSummaryStats:
    """Counters surfaced by the importer for CLI reporting."""

    import_run_id: int = 0
    status: str = "running"
    tnved_categories: int = 0
    business_companies: int = 0
    business_imports: int = 0
    business_import_summaries: int = 0
    region_match_hits: int = 0
    region_match_misses: int = 0
    district_match_hits: int = 0
    district_match_misses: int = 0
    mahalla_match_hits: int = 0
    mahalla_match_misses: int = 0
    error_message: str | None = None


@dataclass
class _GeoResolver:
    """In-memory CERR geography lookup, built once per import run.

    Built from whatever CERR snapshot exists at import time (the most recent
    completed CERR `import_run`). If the chatbot's CERR data is later
    re-imported, the business FKs may technically point at the previous
    snapshot's surrogate ids — that's fine because the FKs are not under a
    cascade and the CERR ids are stable across re-imports anyway.
    """

    regions: dict[str, int] = field(default_factory=dict)
    districts: dict[tuple[int, str], int] = field(default_factory=dict)
    mahallas: dict[tuple[int, str], int] = field(default_factory=dict)
    # Fallback: when we couldn't resolve the district, still try mahalla
    # matches across the whole region.
    mahallas_by_region: dict[tuple[int, str], list[int]] = field(default_factory=dict)

    @classmethod
    def build(cls, session: Session) -> _GeoResolver:
        self = cls()
        for row in session.execute(select(Region.region_id, Region.region_name_cyr)).all():
            key = _normalize_geo_name(row.region_name_cyr)
            if key and key not in self.regions:
                self.regions[key] = row.region_id
        for row in session.execute(
            select(District.district_id, District.region_id, District.district_name_cyr)
        ).all():
            key = _normalize_geo_name(row.district_name_cyr)
            if key:
                composite = (row.region_id, key)
                if composite not in self.districts:
                    self.districts[composite] = row.district_id
        for row in session.execute(
            select(
                Mahalla.mahalla_id,
                Mahalla.district_id,
                Mahalla.region_id,
                Mahalla.mahalla_name_cyr,
            )
        ).all():
            key = _normalize_geo_name(row.mahalla_name_cyr)
            if not key:
                continue
            composite = (row.district_id, key)
            if composite not in self.mahallas:
                self.mahallas[composite] = row.mahalla_id
            self.mahallas_by_region.setdefault((row.region_id, key), []).append(row.mahalla_id)
        return self

    def resolve_region(self, name: Any) -> int | None:
        return self.regions.get(_normalize_geo_name(name))

    def resolve_district(self, region_id: int | None, name: Any) -> int | None:
        if region_id is None:
            return None
        return self.districts.get((region_id, _normalize_geo_name(name)))

    def resolve_mahalla(
        self,
        district_id: int | None,
        region_id: int | None,
        name: Any,
    ) -> int | None:
        key = _normalize_geo_name(name)
        if not key:
            return None
        if district_id is not None:
            hit = self.mahallas.get((district_id, key))
            if hit is not None:
                return hit
        # Mahalla STIR is not globally unique (127 dup groups). If the district
        # didn't resolve, falling back to region-wide match still narrows it
        # down enough to be useful — we take the first hit only when there's
        # exactly one candidate.
        if region_id is not None:
            candidates = self.mahallas_by_region.get((region_id, key))
            if candidates and len(candidates) == 1:
                return candidates[0]
        return None


def import_business(engine: Engine, settings: Settings | None = None) -> BusinessImportSummaryStats:
    """Top-level entrypoint. Loads all four files in a single import_run.

    Phase order matters: tnved_categories first (so business_imports can
    reference them via tnved_code if needed), then business_companies (so
    business_imports + summaries can FK on company_id), then the two TN_VED
    fact tables.
    """
    cfg = settings or get_settings()
    source_dir = Path(getattr(cfg, "business_source_dir", Path("./chatbot/business")))
    stats = BusinessImportSummaryStats()

    # ----- Open import_run row (status='running' first so a crash is visible)
    started_at = datetime.now(UTC)
    run = ImportRun(
        started_at=started_at,
        source_dir=str(source_dir),
        status="running_business",
    )
    with Session(engine, expire_on_commit=False) as s:
        s.add(run)
        s.commit()
        stats.import_run_id = run.import_run_id
        resolver = _GeoResolver.build(s)

    run_id = stats.import_run_id

    try:
        # ----- Phase A: TN_VED dictionary
        tnved_path = source_dir / "TN_VED_description.xls"
        if tnved_path.exists():
            stats.tnved_categories = _load_tnved_categories(engine, tnved_path, run_id=run_id)

        # ----- Phase B: business_companies (OKED.xlsx)
        oked_path = source_dir / "OKED.xlsx"
        if oked_path.exists():
            stats.business_companies = _load_oked_companies(
                engine, oked_path, run_id=run_id, resolver=resolver, stats=stats
            )

        # Build INN → company_id map for FK resolution on imports.
        inn_to_company_id = _load_inn_company_map(engine, run_id=run_id)

        # ----- Phase C: business_imports (TN_VED.xlsx Ойоғи sheet)
        tnved_xlsx = source_dir / "TN_VED.xlsx"
        if tnved_xlsx.exists():
            stats.business_imports = _load_business_imports(
                engine,
                tnved_xlsx,
                run_id=run_id,
                resolver=resolver,
                inn_to_company_id=inn_to_company_id,
                stats=stats,
            )

            # ----- Phase D: business_import_summaries (Мижозлар бўйича sheet)
            stats.business_import_summaries = _load_business_import_summaries(
                engine,
                tnved_xlsx,
                run_id=run_id,
                inn_to_company_id=inn_to_company_id,
            )

    except Exception as exc:  # noqa: BLE001 — annotate, mark failed, re-raise
        with Session(engine, expire_on_commit=False) as s:
            db_run = s.get(ImportRun, run_id)
            if db_run is not None:
                db_run.status = "failed_business"
                db_run.notes = f"{type(exc).__name__}: {exc}"
                s.commit()
        stats.status = "failed_business"
        stats.error_message = f"{type(exc).__name__}: {exc}"
        raise

    # ----- Mark completed_business so views can find the latest run.
    with Session(engine, expire_on_commit=False) as s:
        db_run = s.get(ImportRun, run_id)
        if db_run is not None:
            db_run.completed_at = datetime.now(UTC)
            db_run.status = "completed_business"
            s.commit()
    stats.status = "completed_business"
    return stats


# ---------------------------------------------------------------------------
# Phase A: TN_VED dictionary
# ---------------------------------------------------------------------------


def _load_tnved_categories(engine: Engine, path: Path, *, run_id: int) -> int:
    import xlrd  # type: ignore[import-untyped]

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    inserted = 0
    table = TnvedCategory.__table__
    source_file = path.name
    rows: list[dict[str, Any]] = []
    with engine.begin() as conn:
        # Row 0 is the header in this dictionary file.
        for r in range(1, ws.nrows):
            code = ws.cell_value(r, 0)
            if isinstance(code, float):
                # xlrd returns floats for numeric cells; preserve integer string.
                code = str(int(code))
            elif not isinstance(code, str):
                continue
            code = code.strip()
            if not code:
                continue
            desc = ws.cell_value(r, 1) if ws.ncols > 1 else None
            rows.append(
                {
                    "import_run_id": run_id,
                    "source_file": source_file,
                    "source_row_index": r,
                    "tnved_code": code[:16],
                    "description_ru": desc if isinstance(desc, str) else None,
                    "chapter": code[:2] if len(code) >= 2 else None,
                    "heading": code[:4] if len(code) >= 4 else None,
                }
            )
            if len(rows) >= _BATCH_SIZE:
                conn.execute(table.insert(), rows)
                inserted += len(rows)
                rows.clear()
        if rows:
            conn.execute(table.insert(), rows)
            inserted += len(rows)
    return inserted


# ---------------------------------------------------------------------------
# Phase B: business_companies (OKED.xlsx)
# ---------------------------------------------------------------------------

_OKED_EXPECTED_HEADER = (
    "CLIENT_CODE",
    "INN",
    "COMPANY_NAME",
    "CLIENT_TYPE_NAME",
    "ADDRESS",
    "COUNTRY_NAME",
    "REGION_NAME",
    "DISTRICT_NAME",
    "OKED_NAME_RU",
    "OKED_NAME_UZ",
)


def _load_oked_companies(
    engine: Engine,
    path: Path,
    *,
    run_id: int,
    resolver: _GeoResolver,
    stats: BusinessImportSummaryStats,
) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["main"] if "main" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return 0
        col_idx = _index_columns(header, _OKED_EXPECTED_HEADER)
        table = BusinessCompany.__table__
        source_file = path.name
        inserted = 0
        batch: list[dict[str, Any]] = []
        with engine.begin() as conn:
            for source_row_index, raw in enumerate(rows_iter, start=1):
                if raw is None:
                    continue
                client_code = _cell_str(raw, col_idx["CLIENT_CODE"])
                inn = _cell_str(raw, col_idx["INN"])
                company_name = _cell_str(raw, col_idx["COMPANY_NAME"])
                client_type = _cell_str(raw, col_idx["CLIENT_TYPE_NAME"])
                address_raw = _cell_str(raw, col_idx["ADDRESS"])
                country_name = _cell_str(raw, col_idx["COUNTRY_NAME"])
                region_raw = _cell_str(raw, col_idx["REGION_NAME"])
                district_raw = _cell_str(raw, col_idx["DISTRICT_NAME"])
                oked_ru = _cell_str(raw, col_idx["OKED_NAME_RU"])
                oked_uz = _cell_str(raw, col_idx["OKED_NAME_UZ"])

                region_id = resolver.resolve_region(region_raw)
                district_id = resolver.resolve_district(region_id, district_raw)
                if region_id is not None:
                    stats.region_match_hits += 1
                elif region_raw:
                    stats.region_match_misses += 1
                if district_id is not None:
                    stats.district_match_hits += 1
                elif district_raw:
                    stats.district_match_misses += 1

                batch.append(
                    {
                        "import_run_id": run_id,
                        "source_file": source_file,
                        "source_row_index": source_row_index,
                        "inn": _truncate(inn, 16),
                        "client_code": _truncate(client_code, 16),
                        "company_name": company_name,
                        "client_type_cyr": client_type,
                        "address_raw": address_raw,
                        "country_name_cyr": country_name,
                        "region_name_raw_cyr": region_raw,
                        "district_name_raw_cyr": district_raw,
                        "region_id": region_id,
                        "district_id": district_id,
                        "oked_label_ru": oked_ru,
                        "oked_label_uz": oked_uz,
                    }
                )
                if len(batch) >= _BATCH_SIZE:
                    conn.execute(table.insert(), batch)
                    inserted += len(batch)
                    batch.clear()
            if batch:
                conn.execute(table.insert(), batch)
                inserted += len(batch)
        return inserted
    finally:
        wb.close()


def _load_inn_company_map(engine: Engine, *, run_id: int) -> dict[str, int]:
    """Build {inn: company_id} for the current business import_run.

    Used to populate `business_imports.company_id` and
    `business_import_summaries.company_id`. INN duplicates within the OKED
    feed are observed (different branches of one entity share an INN); we
    keep the first seen company_id to give the chatbot a deterministic
    target — full per-row resolution would require ambiguity handling we
    aren't doing in Phase 2A.
    """
    out: dict[str, int] = {}
    with engine.connect() as conn:
        result = conn.execute(
            text(
                "SELECT inn, company_id FROM business_companies "
                "WHERE import_run_id = :rid AND inn IS NOT NULL"
            ),
            {"rid": run_id},
        )
        for inn, company_id in result:
            if inn and inn not in out:
                out[inn] = company_id
    return out


# ---------------------------------------------------------------------------
# Phase C: business_imports (TN_VED.xlsx "Ойоғи" sheet)
# ---------------------------------------------------------------------------


def _load_business_imports(
    engine: Engine,
    path: Path,
    *,
    run_id: int,
    resolver: _GeoResolver,
    inn_to_company_id: dict[str, int],
    stats: BusinessImportSummaryStats,
) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Ойоғи" not in wb.sheetnames:
            return 0
        ws = wb["Ойоғи"]
        rows_iter = ws.iter_rows(values_only=True)
        # Two-row header: row 0 has parent labels, row 1 has the geographic
        # sub-labels. We use fixed column positions from the layout sample.
        next(rows_iter, None)
        next(rows_iter, None)
        table = BusinessImport.__table__
        source_file = f"{path.name}#Ойоғи"
        inserted = 0
        batch: list[dict[str, Any]] = []
        with engine.begin() as conn:
            for source_row_index, raw in enumerate(rows_iter, start=2):
                if raw is None:
                    continue
                inn = _truncate(_cell_str(raw, 0), 16)
                contract_number = _truncate(_cell_str(raw, 1), 64)
                company_name = _cell_str(raw, 2)
                region_raw = _cell_str(raw, 3)
                district_raw = _cell_str(raw, 4)
                mahalla_raw = _cell_str(raw, 5)
                declaration_number = _truncate(_cell_str(raw, 7), 32)
                declaration_date = _cell_date(raw, 8)
                currency_code = _truncate(_cell_str(raw, 9), 8)
                invoice_value_total = _cell_float(raw, 10)
                origin_country_code = _truncate(_cell_str(raw, 11), 8)
                tnved_code = _truncate(_cell_str(raw, 12), 16)
                invoice_value_item = _cell_float(raw, 13)
                item_description = _cell_str(raw, 14)
                exchange_rate = _cell_float(raw, 15)
                value_usd = _cell_float(raw, 16)

                region_id = resolver.resolve_region(region_raw)
                district_id = resolver.resolve_district(region_id, district_raw)
                mahalla_id = resolver.resolve_mahalla(district_id, region_id, mahalla_raw)
                if mahalla_id is not None:
                    stats.mahalla_match_hits += 1
                elif mahalla_raw:
                    stats.mahalla_match_misses += 1

                tnved_chapter = tnved_code[:2] if tnved_code and len(tnved_code) >= 2 else None
                company_id = inn_to_company_id.get(inn) if inn else None

                batch.append(
                    {
                        "import_run_id": run_id,
                        "source_file": source_file,
                        "source_row_index": source_row_index,
                        "inn": inn,
                        "company_id": company_id,
                        "company_name": company_name,
                        "contract_number": contract_number,
                        "declaration_number": declaration_number,
                        "declaration_date": declaration_date,
                        "region_name_raw_cyr": region_raw,
                        "district_name_raw_cyr": district_raw,
                        "mahalla_name_raw_cyr": mahalla_raw,
                        "region_id": region_id,
                        "district_id": district_id,
                        "mahalla_id": mahalla_id,
                        "tnved_code": tnved_code,
                        "tnved_chapter": tnved_chapter,
                        "currency_code": currency_code,
                        "origin_country_code": origin_country_code,
                        "invoice_value_total": invoice_value_total,
                        "invoice_value_item": invoice_value_item,
                        "value_usd": value_usd,
                        "exchange_rate": exchange_rate,
                        "item_description": item_description,
                        "raw_row_json": None,
                    }
                )
                if len(batch) >= _BATCH_SIZE:
                    conn.execute(table.insert(), batch)
                    inserted += len(batch)
                    batch.clear()
            if batch:
                conn.execute(table.insert(), batch)
                inserted += len(batch)
        return inserted
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Phase D: business_import_summaries (TN_VED.xlsx "Мижозлар бўйича")
# ---------------------------------------------------------------------------


def _load_business_import_summaries(
    engine: Engine,
    path: Path,
    *,
    run_id: int,
    inn_to_company_id: dict[str, int],
) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Мижозлар бўйича" not in wb.sheetnames:
            return 0
        ws = wb["Мижозлар бўйича"]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # row 0: header
        next(rows_iter, None)  # row 1: "Жами" (grand total) — skip
        table = BusinessImportSummary.__table__
        source_file = f"{path.name}#Мижозлар"
        inserted = 0
        batch: list[dict[str, Any]] = []
        with engine.begin() as conn:
            for source_row_index, raw in enumerate(rows_iter, start=2):
                if raw is None:
                    continue
                inn = _truncate(_cell_str(raw, 1), 16)
                company_name = _cell_str(raw, 2)
                if not inn and not company_name:
                    continue
                company_id = inn_to_company_id.get(inn) if inn else None
                batch.append(
                    {
                        "import_run_id": run_id,
                        "source_file": source_file,
                        "source_row_index": source_row_index,
                        "inn": inn,
                        "company_id": company_id,
                        "company_name": company_name,
                        "total_import_usd": _cell_float(raw, 3),
                        "value_machines_usd": _cell_float(raw, 4),
                        "value_chemicals_usd": _cell_float(raw, 5),
                        "value_misc_goods_usd": _cell_float(raw, 6),
                        "value_industrial_usd": _cell_float(raw, 7),
                        "value_animal_oils_usd": _cell_float(raw, 8),
                        "value_non_food_raw_usd": _cell_float(raw, 9),
                        "value_building_mats_usd": _cell_float(raw, 10),
                        "value_food_products_usd": _cell_float(raw, 11),
                        "value_fruit_veg_usd": _cell_float(raw, 12),
                        "value_mineral_fuel_usd": _cell_float(raw, 13),
                        "value_beverages_tobacco_usd": _cell_float(raw, 14),
                        "value_other_goods_usd": _cell_float(raw, 15),
                    }
                )
                if len(batch) >= _BATCH_SIZE:
                    conn.execute(table.insert(), batch)
                    inserted += len(batch)
                    batch.clear()
            if batch:
                conn.execute(table.insert(), batch)
                inserted += len(batch)
        return inserted
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _index_columns(header: Iterable[Any], expected: tuple[str, ...]) -> dict[str, int]:
    """Map each expected column name to its index in the actual header.

    The OKED file has the expected layout, but column order can differ in
    other exports. Missing columns map to a sentinel -1; cell helpers treat
    -1 as "no value".
    """
    by_name: dict[str, int] = {}
    for idx, value in enumerate(header):
        if isinstance(value, str):
            by_name[value.strip()] = idx
    return {name: by_name.get(name, -1) for name in expected}


def _cell_str(row: Any, idx: int) -> str | None:
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def _cell_float(row: Any, idx: int) -> float | None:
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(" ", "").replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _cell_date(row: Any, idx: int) -> date | None:
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _truncate(value: str | None, max_len: int) -> str | None:
    if value is None:
        return None
    return value[:max_len]


__all__ = [
    "BusinessImportSummaryStats",
    "import_business",
]
