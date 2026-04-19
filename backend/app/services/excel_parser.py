"""Parses uploaded Excel files using Claude AI for intelligent extraction.

Converts Excel sheets to text, sends to Claude with a structured prompt,
and returns parsed financial data (codes, absolutes, ratios).
Falls back to rule-based parsing if Claude is unavailable.
"""
from __future__ import annotations
import json
import logging
import re
from io import BytesIO
from typing import Literal

from openpyxl import load_workbook

from ..config import get_settings

log = logging.getLogger(__name__)

CODE_RE = re.compile(r"^\d{3}$")
FormKind = Literal["balance", "pnl"]

# ---------------------------------------------------------------------------
# Excel → text conversion
# ---------------------------------------------------------------------------

def _excel_to_text(blob: bytes) -> str:
    """Convert all sheets of an Excel file to readable text for Claude."""
    wb = load_workbook(BytesIO(blob), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(v) if v is not None else "" for v in row]
            # Skip completely empty rows
            if not any(c.strip() for c in cells):
                continue
            parts.append("\t".join(cells))
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Claude-based parsing
# ---------------------------------------------------------------------------

CLAUDE_PARSE_PROMPT = """You are a financial document parser. You receive the text content of an Excel file
(Uzbek accounting Form 1 — Balance Sheet, or Form 2 — Profit & Loss statement, or any other financial report).

Your task: extract ALL financial data and return a structured JSON.

For Uzbek standard forms, extract the 3-digit line codes (010, 020, 030, etc.) and their values.
For non-standard formats, identify the financial metrics by their meaning.

Return STRICTLY valid JSON (no markdown, no explanation) with this structure:

{
  "codes": {
    "010": 1234567.00,
    "020": 987654.00
  },
  "absolutes": {
    "revenue": null,
    "grossProfit": null,
    "operatingProfit": null,
    "netProfit": null,
    "totalAssets": null,
    "equity": null,
    "totalLiabilities": null,
    "totalDebt": null,
    "currentAssets": null,
    "currentLiabilities": null,
    "inventory": null,
    "receivables": null,
    "cash": null,
    "interestExpense": null
  },
  "ratios": {
    "grossMargin": null,
    "operatingMargin": null,
    "netMargin": null,
    "roa": null,
    "roe": null,
    "currentRatio": null,
    "quickRatio": null,
    "debtToEquity": null,
    "debtToAssets": null,
    "assetTurnover": null,
    "inventoryTurn": null,
    "interestCoverage": null
  }
}

Rules:
- "codes": extract ALL 3-digit codes found in the document with their numeric values. If the document doesn't use standard codes, leave this as empty {}.
- "absolutes": fill in the absolute financial values you can identify from the document. Use null for values you cannot find.
- "ratios": CALCULATE all ratios you can from the absolutes. Use null if you lack the data to calculate.
  - grossMargin = grossProfit / revenue
  - operatingMargin = operatingProfit / revenue
  - netMargin = netProfit / revenue
  - roa = netProfit / totalAssets
  - roe = netProfit / equity
  - currentRatio = currentAssets / currentLiabilities
  - quickRatio = (currentAssets - inventory) / currentLiabilities
  - debtToEquity = totalDebt / equity
  - debtToAssets = totalDebt / totalAssets
  - assetTurnover = revenue / totalAssets
  - inventoryTurn = costOfGoodsSold / inventory
  - interestCoverage = operatingProfit / interestExpense
- For Uzbek Form 1 (Balance): revenue and profit fields come from Form 2, so set them to null. Focus on assets, equity, liabilities.
- For Uzbek Form 2 (PnL): asset/equity fields come from Form 1, so set them to null. Focus on revenue, costs, profit.
- All numeric values should be plain numbers (no formatting, no spaces, no currency symbols).
- If a value is clearly 0, use 0 (not null). Use null only when the data is not available.
"""


def _parse_with_claude(blob: bytes, kind: str) -> dict:
    """Send Excel content to Claude for intelligent parsing."""
    from anthropic import Anthropic

    settings = get_settings()
    if not settings.anthropic_api_key_clean:
        raise RuntimeError("ANTHROPIC_API_KEY is not configured")

    text_content = _excel_to_text(blob)

    client = Anthropic(api_key=settings.anthropic_api_key_clean)
    resp = client.messages.create(
        model=settings.anthropic_model_clean,
        max_tokens=2000,
        system=CLAUDE_PARSE_PROMPT,
        messages=[{
            "role": "user",
            "content": f"Document type: {kind} ({'Balance Sheet' if kind == 'balance' else 'Profit & Loss'})\n\n{text_content}"
        }],
    )

    text_parts = [b.text for b in resp.content if getattr(b, "type", None) == "text"]
    raw = "\n".join(text_parts).strip()

    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].lstrip()

    parsed = json.loads(raw)

    usage = getattr(resp, "usage", None)
    log.info(
        "Claude parsed %s: %d codes, tokens in=%d out=%d",
        kind,
        len(parsed.get("codes", {})),
        getattr(usage, "input_tokens", 0) if usage else 0,
        getattr(usage, "output_tokens", 0) if usage else 0,
    )

    return parsed


# ---------------------------------------------------------------------------
# Rule-based fallback (original parser)
# ---------------------------------------------------------------------------

def _as_num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "x", "X", "-", "—"):
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _row_code_and_values(row):
    code_idx = None
    for j, v in enumerate(row):
        if v is None:
            continue
        if CODE_RE.match(str(v).strip()):
            code_idx = j
            break
    if code_idx is None:
        return None, []
    return str(row[code_idx]).strip(), [_as_num(v) for v in row[code_idx + 1:]]


def _pick_values_sheet(wb):
    if "list02" in wb.sheetnames:
        return wb["list02"]
    return wb.active


def _parse_balance_rules(blob: bytes) -> dict[str, float]:
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = _pick_values_sheet(wb)
    out: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = _row_code_and_values(row)
        if not code:
            continue
        nums = [v for v in vals if v is not None]
        if not nums:
            continue
        out[code] = nums[1] if len(nums) >= 2 else nums[0]
    return out


def _parse_pnl_rules(blob: bytes) -> dict[str, float]:
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = _pick_values_sheet(wb)
    out: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = _row_code_and_values(row)
        if not code:
            continue
        padded = (vals + [None] * 4)[:4]
        cur_income, cur_expense = padded[2], padded[3]
        if cur_income is not None:
            out[code] = cur_income
        elif cur_expense is not None:
            out[code] = cur_expense
    return out


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse(kind: FormKind, blob: bytes) -> dict:
    """Parse Excel file — uses Claude AI, falls back to rule-based parser."""
    settings = get_settings()

    # Try Claude first if API key is available
    if settings.anthropic_api_key:
        try:
            result = _parse_with_claude(blob, kind)
            # Return in the format expected by the route:
            # { "codes": {...}, "computed": { "absolutes": {...}, "ratios": {...} } }
            return {
                "codes": result.get("codes", {}),
                "computed": {
                    "absolutes": result.get("absolutes", {}),
                    "ratios": result.get("ratios", {}),
                },
            }
        except Exception as e:
            log.warning("Claude parsing failed, falling back to rules: %s", e)

    # Fallback to rule-based parsing
    if kind == "balance":
        codes = _parse_balance_rules(blob)
    elif kind == "pnl":
        codes = _parse_pnl_rules(blob)
    else:
        raise ValueError(f"Unknown form kind: {kind}")

    return codes
