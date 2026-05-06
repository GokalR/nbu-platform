"""SME Profile data layer: questions catalogue + Excel reference loaders.

- Questions: hardcoded sphere catalogue (was originally Excel-backed in the
  standalone platform; inlined here so we don't bundle questions.xlsx).
- Reference Excel files (read-only):
    * ruyxat.xlsx     → INN/PINFL → client info lookup (was «Руйхат-2 (1).xlsx»)
    * tuman_mfy.xlsx  → viloyat → tuman → MFY cascade (was «Туман + МФЙ.xlsx»)
  Both live in backend/data/sme_profile/. ruyxat.xlsx is gitignored
  (sensitive bank client data); the loader returns gracefully empty when
  missing. Legacy filenames are still recognised so admins can drop the
  files in unrenamed.

  On Railway, ruyxat.xlsx isn't bundled in the image. Set these env vars
  to make the loader fetch it from R2 on first lookup:
    R2_ENDPOINT             https://<account>.r2.cloudflarestorage.com
    R2_ACCESS_KEY_ID        (from R2 API token)
    R2_SECRET_ACCESS_KEY    (from R2 API token)
    R2_BUCKET               nbu-platform-data
    R2_RUYXAT_KEY           ruyxat.xlsx           (optional, default)

Implementation uses openpyxl (already a dependency) — pandas is not
required, which avoids the silent ImportError trap that left lookups
returning «not found» on machines where pandas wasn't installed.
"""

from __future__ import annotations

import functools
import math
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

# ── Paths ──────────────────────────────────────────────────────────────────
_BACKEND_DIR = Path(__file__).parent.parent.parent  # backend/
_DATA_DIR = _BACKEND_DIR / "data" / "sme_profile"

# Clean ASCII names first; legacy Cyrillic names as fallback so admins can
# drop the originals straight from the data team without renaming.
_RUYXAT_FILES = ("ruyxat.xlsx", "Руйхат-2 (1).xlsx", "Руйхат-2.xlsx")
_TUMAN_MFY_FILES = ("tuman_mfy.xlsx", "Туман + МФЙ.xlsx")


def _find(filenames: tuple[str, ...] | str) -> Optional[Path]:
    """Locate a reference Excel file. Honours SME_PROFILE_DATA_DIR env var
    so Railway/R2-mounted volumes can override the default."""
    if isinstance(filenames, str):
        filenames = (filenames,)
    candidates: List[Path] = []
    override = os.getenv("SME_PROFILE_DATA_DIR", "").strip()
    bases = []
    if override:
        bases.append(Path(override))
    bases.append(_DATA_DIR)
    for base in bases:
        for fname in filenames:
            candidates.append(base / fname)
    for p in candidates:
        if p.exists():
            return p
    return None


# ── Value cleaner (handles NaN, "nan", 0, "0") ────────────────────────────
def _clean(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "0", ""):
        return ""
    return s


def _format_num(val: Any) -> str:
    """Format a numeric Excel cell as «123 456 789». Returns '' for empties."""
    s = _clean(val)
    if not s:
        return ""
    try:
        return "{:,.0f}".format(float(s)).replace(",", " ")
    except (ValueError, TypeError):
        return s


# ── Address cascade (вилоят → туман → МФЙ) ────────────────────────────────
@functools.lru_cache(maxsize=1)
def load_address_data() -> Dict[str, Dict[str, List[str]]]:
    """Returns {viloyat: {tuman: [mfy_list]}} from tuman_mfy.xlsx.

    Sheet layout: rows 1–2 are blank/title, row 3 is the header, data
    starts row 4 with merged cells in columns C (viloyat) and D (tuman).
    We forward-fill the merged values manually because openpyxl's
    read_only mode reports merged-cell continuations as None.
    """
    path = _find(_TUMAN_MFY_FILES)
    if not path:
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        # Sheet name is «МФЙ» but using index 0 avoids any name-encoding flakiness.
        ws = wb.worksheets[0]

        result: Dict[str, Dict[str, List[str]]] = {}
        last_v = ""
        last_t = ""
        for row in ws.iter_rows(min_row=4, values_only=True):
            v_raw = row[2] if len(row) > 2 else None
            t_raw = row[3] if len(row) > 3 else None
            m_raw = row[4] if len(row) > 4 else None
            v = _clean(v_raw)
            t = _clean(t_raw)
            m = _clean(m_raw)
            if v:
                last_v = v
                # New viloyat starts → reset tuman so a stale carry-over
                # from the previous viloyat block doesn't leak in.
                if not t:
                    last_t = ""
            if t:
                last_t = t
            if m and last_v and last_t:
                result.setdefault(last_v, {}).setdefault(last_t, []).append(m)
        wb.close()
        return result
    except Exception as exc:
        print(f"[sme_profile.address] load failed: {exc}")
        return {}


# ── R2 fetch (production) ─────────────────────────────────────────────────
def _ensure_ruyxat_from_r2() -> Optional[Path]:
    """If ruyxat.xlsx isn't on disk and R2 creds are set, download it.

    Idempotent: skips if the file already exists locally. Returns the
    path to the downloaded file, or None if R2 isn't configured / the
    download failed. The caller falls back to the regular `_find()`
    search after this.
    """
    target = _DATA_DIR / "ruyxat.xlsx"
    if target.exists():
        return target

    endpoint = os.getenv("R2_ENDPOINT", "").strip()
    access_key = os.getenv("R2_ACCESS_KEY_ID", "").strip()
    secret_key = os.getenv("R2_SECRET_ACCESS_KEY", "").strip()
    bucket = os.getenv("R2_BUCKET", "").strip()
    key = os.getenv("R2_RUYXAT_KEY", "ruyxat.xlsx").strip()

    if not all([endpoint, access_key, secret_key, bucket]):
        return None

    try:
        import boto3
        target.parent.mkdir(parents=True, exist_ok=True)
        client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name="auto",
        )
        # Download to a temp path then rename, so a partial download
        # never leaves a half-written file the loader could open.
        tmp = target.with_suffix(".xlsx.partial")
        client.download_file(bucket, key, str(tmp))
        tmp.replace(target)
        size_mb = target.stat().st_size / 1024 / 1024
        print(f"[sme_profile] downloaded ruyxat.xlsx from R2 ({size_mb:.1f} MB)")
        return target
    except Exception as exc:
        print(f"[sme_profile] R2 download failed: {exc}")
        return None


# ── Client lookup (Руйхат-2 → ruyxat.xlsx) ────────────────────────────────
@functools.lru_cache(maxsize=1)
def load_client_db() -> Dict[str, Dict[str, str]]:
    """Returns {INN: ClientInfo fields} from ruyxat.xlsx.

    File layout: row 1 is the header, row 2 is empty, data starts row 3.
    Two columns are duplicated in the header (BRANCH_ID, OKED) — for
    'activity_type' we use the second OKED column to match the original
    pandas-based behaviour (`OKED.1`).

    Production: if the file isn't on disk and R2_* env vars are set,
    fetch it from R2 first.
    """
    _ensure_ruyxat_from_r2()
    path = _find(_RUYXAT_FILES)
    if not path:
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        # First-occurrence map for unique columns; explicit second occurrence
        # for OKED (activity_type uses OKED.1 in the original pandas code).
        col_idx: Dict[str, int] = {}
        for i, name in enumerate(header):
            if isinstance(name, str) and name and name not in col_idx:
                col_idx[name] = i
        oked_positions = [i for i, n in enumerate(header) if n == "OKED"]
        oked_col = oked_positions[1] if len(oked_positions) >= 2 else (
            oked_positions[0] if oked_positions else None
        )

        def cell(row: tuple, name: str) -> Any:
            i = col_idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        result: Dict[str, Dict[str, str]] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            inn = _clean(cell(row, "INN"))
            if not inn:
                continue
            if inn.endswith(".0"):
                inn = inn[:-2]
            if not inn:
                continue

            phone = _clean(cell(row, "PHONE")) or _clean(cell(row, "MOBILE_PHONE"))

            shareholders = _clean(cell(row, "CNT_FL"))
            if shareholders.endswith(".0"):
                shareholders = shareholders[:-2]

            activity = ""
            if oked_col is not None and oked_col < len(row):
                activity = _clean(row[oked_col])

            result[inn] = {
                "company_name":       _clean(cell(row, "CLIENT_NAME")),
                "director":           _clean(cell(row, "DIRECTOR_NAME")),
                "reg_address":        _clean(cell(row, "ADDRESS")),
                "phone":              phone,
                "turnover_debit":     _format_num(cell(row, "TURNOVER_DEBIT")),
                "turnover_credit":    _format_num(cell(row, "TURNOVER_CREDIT")),
                "turnover_all":       _format_num(cell(row, "TURNOVER_ALL")),
                "shareholders_count": shareholders,
                "accountant":         _clean(cell(row, "ACCOUNTER_CHIEF_NAME")),
                "activity_type":      activity,
                "sal_sum":            _format_num(cell(row, "SAL_SUM")),
            }
        wb.close()
        return result
    except Exception as exc:
        print(f"[sme_profile.client_db] load failed: {exc}")
        return {}


# ── Question catalogue (sphere-by-sphere) ─────────────────────────────────
# Tuple shape:
# (cat_id, cat_ru, cat_uz, icon, q_id, q_ru, q_uz, q_type, opts_ru, opts_uz)
#
# Icons map to Google Material Symbols (the platform's <AppIcon> uses
# material-symbols-outlined). Keep this list in sync with the steps UI.
_ROWS: List[tuple] = [
    # ── Торговое направление / Савдо йўналиши ──────────────────────────
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "employees_count", "Количество действующих сотрудников", "Фаол ходимлар сони",
     "number", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "activity_type", "Вид деятельности", "Фаолият тури",
     "select",
     "Розничная торговля;Оптовая торговля;Оба вида",
     "Чакана савдо;Улгуржи савдо;Иккаласи ҳам"),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "related_companies_count", "Количество связанных компаний", "Боғлиқ компаниялар сони",
     "number", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "purchased_goods", "Наименование закупаемых товаров", "Харид қилинадиган товарлар номи",
     "textarea", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "purchase_address", "Адрес закупки товаров", "Товар харид манзили",
     "text", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "financial_problems", "Финансовые затруднения", "Молиявий қийинчиликлар",
     "checkbox",
     "Недостаток оборотных средств;Высокие налоги;Проблемы с кредитованием;Инфляция;Колебания курса валют",
     "Айланма маблағлар етишмаслиги;Юқори солиқлар;Кредитлашдаги муаммолар;Инфляция;Валюта курси ўзгариши"),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "problems_description", "Опишите проблемы подробнее", "Муаммоларни батафсил тавсифланг",
     "textarea", "", ""),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "solutions", "Предложения по решению проблем", "Муаммоларни ҳал қилиш бўйича таклифлар",
     "checkbox",
     "Снижение налоговой нагрузки;Льготное кредитование;Субсидии;Обучение персонала;Цифровизация",
     "Солиқ юкини камайтириш;Имтиёзли кредитлаш;Субсидиялар;Ходимларни ўқитиш;Рақамлаштириш"),
    ("trade", "Торговое направление", "Савдо йўналиши", "shopping_bag",
     "solutions_comment", "Комментарии к предложениям", "Таклифларга изоҳ",
     "textarea", "", ""),

    # ── Медицинское направление / Тиббиёт йўналиши ────────────────────
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "employees_count", "Количество действующих сотрудников", "Фаол ходимлар сони",
     "number", "", ""),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "activity_type", "Вид медицинской деятельности", "Тиббий фаолият тури",
     "select",
     "Клиника;Аптека;Лаборатория;Стоматология;Другое",
     "Клиника;Дорихона;Лаборатория;Стоматология;Бошқа"),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "license_available", "Наличие лицензии", "Лицензия мавжудлиги",
     "radio",
     "Да;Нет;В процессе оформления",
     "Ҳа;Йўқ;Расмийлаштириш жараёнида"),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "equipment_problems", "Проблемы с оборудованием", "Жиҳозлар билан муаммолар",
     "checkbox",
     "Устаревшее оборудование;Нехватка запчастей;Высокая стоимость;Сложность обслуживания",
     "Эскирган жиҳозлар;Эҳтиёт қисмлар етишмаслиги;Юқори нарх;Техник хизмат қийинчилиги"),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "financial_problems", "Финансовые затруднения", "Молиявий қийинчиликлар",
     "checkbox",
     "Недостаток оборотных средств;Высокие налоги;Проблемы с кредитованием;Инфляция",
     "Айланма маблағлар етишмаслиги;Юқори солиқлар;Кредитлашдаги муаммолар;Инфляция"),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "problems_description", "Опишите проблемы подробнее", "Муаммоларни батафсил тавсифланг",
     "textarea", "", ""),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "solutions", "Предложения по решению проблем", "Муаммоларни ҳал қилиш бўйича таклифлар",
     "checkbox",
     "Снижение налоговой нагрузки;Льготное кредитование;Субсидии;Обучение персонала",
     "Солиқ юкини камайтириш;Имтиёзли кредитлаш;Субсидиялар;Ходимларни ўқитиш"),
    ("medical", "Медицинское направление", "Тиббиёт йўналиши", "medical_services",
     "solutions_comment", "Комментарии к предложениям", "Таклифларга изоҳ",
     "textarea", "", ""),

    # ── Розничная торговля / Чакана савдо ─────────────────────────────
    ("retail", "Розничная торговля", "Чакана савдо", "shopping_cart",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("retail", "Розничная торговля", "Чакана савдо", "shopping_cart",
     "employees_count", "Количество сотрудников", "Ходимлар сони",
     "number", "", ""),
    ("retail", "Розничная торговля", "Чакана савдо", "shopping_cart",
     "product_type", "Вид товара", "Товар тури",
     "select",
     "Продукты питания;Одежда/обувь;Электроника;Стройматериалы;Другое",
     "Озиқ-овқат;Кийим/пойафзал;Электроника;Қурилиш материаллари;Бошқа"),
    ("retail", "Розничная торговля", "Чакана савдо", "shopping_cart",
     "problems_description", "Основные проблемы бизнеса", "Бизнесдаги асосий муаммолар",
     "textarea", "", ""),

    # ── Услуги / Хизматлар ─────────────────────────────────────────────
    ("services", "Услуги", "Хизматлар", "business_center",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("services", "Услуги", "Хизматлар", "business_center",
     "employees_count", "Количество сотрудников", "Ходимлар сони",
     "number", "", ""),
    ("services", "Услуги", "Хизматлар", "business_center",
     "service_type", "Вид услуг", "Хизмат тури",
     "text", "", ""),
    ("services", "Услуги", "Хизматлар", "business_center",
     "problems_description", "Основные проблемы", "Асосий муаммолар",
     "textarea", "", ""),

    # ── Общественное питание / Жамоат овқатланиши ─────────────────────
    ("food", "Общественное питание", "Жамоат овқатланиши", "restaurant",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("food", "Общественное питание", "Жамоат овқатланиши", "restaurant",
     "employees_count", "Количество сотрудников", "Ходимлар сони",
     "number", "", ""),
    ("food", "Общественное питание", "Жамоат овқатланиши", "restaurant",
     "cuisine_type", "Вид кухни", "Ошхона тури",
     "select",
     "Узбекская;Европейская;Фастфуд;Смешанная;Другое",
     "Ўзбек;Европа;Fastfood;Аралаш;Бошқа"),
    ("food", "Общественное питание", "Жамоат овқатланиши", "restaurant",
     "problems_description", "Основные проблемы", "Асосий муаммолар",
     "textarea", "", ""),

    # ── Строительство / Қурилиш ────────────────────────────────────────
    ("construction", "Строительство", "Қурилиш", "engineering",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("construction", "Строительство", "Қурилиш", "engineering",
     "employees_count", "Количество сотрудников", "Ходимлар сони",
     "number", "", ""),
    ("construction", "Строительство", "Қурилиш", "engineering",
     "construction_type", "Вид строительства", "Қурилиш тури",
     "select",
     "Жилое строительство;Коммерческое;Промышленное;Ремонт",
     "Турар-жой қурилиши;Тижорат;Саноат;Таъмир"),
    ("construction", "Строительство", "Қурилиш", "engineering",
     "problems_description", "Основные проблемы", "Асосий муаммолар",
     "textarea", "", ""),

    # ── Сельское хозяйство / Қишлоқ хўжалиги ─────────────────────────
    ("agriculture", "Сельское хозяйство", "Қишлоқ хўжалиги", "agriculture",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("agriculture", "Сельское хозяйство", "Қишлоқ хўжалиги", "agriculture",
     "land_area", "Площадь земли (га)", "Ер майдони (га)",
     "number", "", ""),
    ("agriculture", "Сельское хозяйство", "Қишлоқ хўжалиги", "agriculture",
     "crop_type", "Вид культуры / животноводства", "Экин / чорвачилик тури",
     "text", "", ""),
    ("agriculture", "Сельское хозяйство", "Қишлоқ хўжалиги", "agriculture",
     "problems_description", "Основные проблемы", "Асосий муаммолар",
     "textarea", "", ""),

    # ── Логистика / Логистика ──────────────────────────────────────────
    ("logistics", "Логистика", "Логистика", "local_shipping",
     "activity_address", "Адрес деятельности", "Фаолият манзили",
     "address_cascade", "", ""),
    ("logistics", "Логистика", "Логистика", "local_shipping",
     "fleet_size", "Количество транспортных средств", "Транспорт воситалари сони",
     "number", "", ""),
    ("logistics", "Логистика", "Логистика", "local_shipping",
     "transport_type", "Вид транспорта", "Транспорт тури",
     "select",
     "Грузовой;Пассажирский;Смешанный",
     "Юк;Йўловчи;Аралаш"),
    ("logistics", "Логистика", "Логистика", "local_shipping",
     "problems_description", "Основные проблемы", "Асосий муаммолар",
     "textarea", "", ""),
]


def load_questions() -> List[Dict[str, Any]]:
    """Group _ROWS into the categories shape the frontend expects."""
    categories: Dict[str, Any] = {}
    for (
        cat_id, cat_ru, cat_uz, icon,
        q_id, q_ru, q_uz, q_type, opts_ru, opts_uz,
    ) in _ROWS:
        if cat_id not in categories:
            categories[cat_id] = {
                "id": cat_id,
                "name": {"ru": cat_ru, "uz": cat_uz},
                "icon": icon,
                "questions": [],
            }
        categories[cat_id]["questions"].append({
            "id": q_id,
            "text": {"ru": q_ru, "uz": q_uz},
            "type": q_type,
            "options": {
                "ru": [o.strip() for o in opts_ru.split(";") if o.strip()],
                "uz": [o.strip() for o in opts_uz.split(";") if o.strip()],
            },
        })
    return list(categories.values())
