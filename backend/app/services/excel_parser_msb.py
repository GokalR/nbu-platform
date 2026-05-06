"""Parser for the standard Uzbekistan SME tax forms:
  • Form №1 — Бухгалтерский баланс (Balance Sheet)
  • Form №2 — Отчёт о финансовых результатах (P&L)

Both forms are issued by Davlat soliq qo'mitasi and have identical
structure across every SME — extraction is by row code, not by header
text or column position guessing. Values are returned in soums (UZS),
even though the source files are in 'тыс. сум' (thousand soums) — we
multiply by 1000 on the way out.
"""
from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl

log = logging.getLogger(__name__)


# ---------- Form №1 (Balance) — row codes we care about ----------
BALANCE_ROW_NAMES = {
    "012": "fixedAssetsNetBook",       # Остаточная стоимость ОС
    "130": "nonCurrentAssets",         # Итого по разделу I
    "140": "inventory",                # ТМЗ всего
    "210": "receivables",              # Дебиторы всего
    "320": "cash",                     # Денежные средства всего
    "390": "currentAssets",            # Итого по разделу II
    "400": "totalAssets",              # БАЛАНС актива
    "410": "charterCapital",           # Уставный капитал
    "450": "retainedEarnings",         # Нераспределённая прибыль
    "480": "equity",                   # Источники собственных средств
    "490": "longTermLiabilities",      # Долгосрочные обязательства
    "600": "currentLiabilities",       # Текущие обязательства
    "770": "totalLiabilities",         # Раздел II пассива
    "780": "totalLiabilitiesAndEquity",# БАЛАНС пассива
}

# ---------- Form №2 (P&L) — row codes we care about ----------
PNL_ROW_NAMES = {
    "010": "revenue",          # income side
    "020": "cogs",             # expense side
    "030": "grossProfit",      # signed
    "040": "periodExpenses",   # expense side
    "100": "operatingProfit",  # signed
    "220": "pretaxProfitOps",  # signed
    "240": "pretaxProfit",     # signed
    "270": "netProfit",        # signed
}


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _norm_code(v: Any) -> str | None:
    """Row codes are 3-digit strings like '010', '040', '770'."""
    if v is None:
        return None
    s = str(v).strip()
    if s.isdigit() and len(s) == 3:
        return s
    return None


def parse_balance(file_bytes: bytes) -> dict[str, Any]:
    """Parse Form №1. Returns {field_name: amount_uzs} for known rows.

    Tries every column 0..9 to find row codes; values come from columns
    immediately to the right (begin / end period). We use the END period.
    All values are converted from тыс. сум → сум by ×1000.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    raw: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        # Find the row-code cell (any of the first 5 columns)
        for col_idx in range(min(5, len(row))):
            code = _norm_code(row[col_idx])
            if code and code in BALANCE_ROW_NAMES:
                # Value cells are at code_col+1 (begin) and code_col+2 (end).
                # We use END period.
                end_val = row[col_idx + 2] if len(row) > col_idx + 2 else None
                if end_val in (None, "", " "):
                    end_val = row[col_idx + 1] if len(row) > col_idx + 1 else None
                raw[code] = _to_float(end_val)
                break

    # Multiply by 1000 (form is in тыс.сум) and rename to friendly fields
    out: dict[str, float] = {}
    for code, name in BALANCE_ROW_NAMES.items():
        out[name] = raw.get(code, 0.0) * 1000

    # Sanity: totalLiabilities should equal longTerm + current. If 770 was
    # missing but the other two are present, derive it.
    if not out["totalLiabilities"] and (out["longTermLiabilities"] or out["currentLiabilities"]):
        out["totalLiabilities"] = out["longTermLiabilities"] + out["currentLiabilities"]

    out["_raw"] = raw  # keep raw codes for admin debugging
    return out


def parse_pnl(file_bytes: bytes) -> dict[str, Any]:
    """Parse Form №2. The P&L has separate income and expense columns —
    we read both and pick whichever side has a value (or the difference
    for signed rows like grossProfit/operatingProfit/netProfit).

    Layout (data rows):
       col 0: №
       col 1: name (uz)
       col 2: blank
       col 3: code (e.g. '010')
       col 4: prior_year_income
       col 5: prior_year_expense
       col 6: current_period_income
       col 7: current_period_expense
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    raw: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        code = _norm_code(row[3])
        if code and code in PNL_ROW_NAMES:
            raw[code] = {
                "income": _to_float(row[6]),
                "expense": _to_float(row[7]),
            }

    out: dict[str, float] = {}
    for code, name in PNL_ROW_NAMES.items():
        cell = raw.get(code, {"income": 0, "expense": 0})
        # Signed rows: profit = income - expense
        out[name] = (cell["income"] - cell["expense"]) * 1000

    # For pure-income/expense rows we want the raw side, not the diff.
    # Re-extract these:
    if "010" in raw:
        out["revenue"] = raw["010"]["income"] * 1000
    if "020" in raw:
        out["cogs"] = raw["020"]["expense"] * 1000
    if "040" in raw:
        out["periodExpenses"] = raw["040"]["expense"] * 1000

    out["_raw"] = raw
    return out
