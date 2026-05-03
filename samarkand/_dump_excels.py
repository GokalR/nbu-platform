"""Dump every cell of both Samarkand Excels as JSON for inspection."""
import json, os, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
files = [
    "Самарқанд вил ПРОМП-вилоят.xlsx",
    "ПРОМТ қўшимча Самарқанд вилоят (3).xlsx",
]
out = {}
for fn in files:
    path = os.path.join(HERE, fn)
    wb = load_workbook(path, data_only=True)
    book = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                rows.append([None if c is None else (c if isinstance(c, (int, float, bool)) else str(c)) for c in row])
        book[sn] = {"max_row": ws.max_row, "max_col": ws.max_column, "rows": rows}
    out[fn] = book

dst = os.path.join(HERE, "_dump.json")
with open(dst, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("WROTE", dst)
print("FILES:", list(out.keys()))
for fn, book in out.items():
    print(f"\n=== {fn} ===")
    for sn, sd in book.items():
        print(f"  sheet '{sn}': {sd['max_row']} rows × {sd['max_col']} cols, kept {len(sd['rows'])} non-empty rows")
