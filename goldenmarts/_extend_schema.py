"""One-shot schema extension for the Golden Mart Excel files.

Adds:
  - golden_mart_city.xlsx, sheet "Города и туманы":
      * 17 detailed age brackets in section 1 (after "Население 65+ лет")
      * "Рождений (всего)" + "Смертей (всего)" in section 11
  - golden_mart_region.xlsx, sheet "Регионы":
      * New section 20 "Темпы реального роста по секторам" at the end

Run from project root:
    PYTHONIOENCODING=utf-8 python goldenmarts/_extend_schema.py

Idempotent — checks if the new rows are already present before inserting.
"""
from __future__ import annotations
import os
import openpyxl as o
from copy import copy

HERE = os.path.dirname(__file__)


def find_row_with_text(ws, col, text):
    """Return 1-based row index of first row in `col` whose cell starts with `text`."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip().startswith(text):
            return r
    return None


def copy_row_style(ws, src_row, dst_row, max_col):
    """Copy style from src_row to dst_row across columns 1..max_col."""
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


# ────────────────────────────────────────────────────────────────────────
# 1. City Excel — extend section 1 (age brackets) + section 11 (counts)
# ────────────────────────────────────────────────────────────────────────
def extend_city():
    path = os.path.join(HERE, 'golden_mart_city.xlsx')
    wb = o.load_workbook(path)
    ws = wb['Города и туманы']

    LABEL_COL = 2   # Cyrillic data sits in column B (column 1 is empty)
    UNIT_COL  = 3
    EX_COL    = 4

    age_brackets = [
        ('Население 0–2 лет',   'чел.', 22000),
        ('Население 3–5 лет',   'чел.', 19000),
        ('Население 6–7 лет',   'чел.', 11000),
        ('Население 8–15 лет',  'чел.', 43000),
        ('Население 16–17 лет', 'чел.', 11000),
        ('Население 18–19 лет', 'чел.',  9800),
        ('Население 20–24 лет', 'чел.', 21000),
        ('Население 25–29 лет', 'чел.', 22000),
        ('Население 30–34 лет', 'чел.', 26000),
        ('Население 35–39 лет', 'чел.', 24000),
        ('Население 40–49 лет', 'чел.', 39000),
        ('Население 50–59 лет', 'чел.', 27000),
        ('Население 60–69 лет', 'чел.', 22000),
        ('Население 70–74 лет', 'чел.',  6300),
        ('Население 75–79 лет', 'чел.',  3500),
        ('Население 80–84 лет', 'чел.',  1300),
        ('Население 85+ лет',   'чел.',   880),
    ]

    if find_row_with_text(ws, LABEL_COL, 'Население 0–2 лет'):
        print('  city: age brackets already present, skipping')
    else:
        anchor = find_row_with_text(ws, LABEL_COL, 'Население 65+ лет')
        if not anchor:
            raise RuntimeError('Could not locate "Население 65+ лет" anchor row')
        ws.insert_rows(anchor + 1, amount=len(age_brackets))
        for i, (label, unit, example) in enumerate(age_brackets):
            r = anchor + 1 + i
            ws.cell(row=r, column=LABEL_COL).value = label
            ws.cell(row=r, column=UNIT_COL).value = unit
            ws.cell(row=r, column=EX_COL).value = example
            copy_row_style(ws, anchor, r, ws.max_column)
        print(f'  city: inserted {len(age_brackets)} age-bracket rows after row {anchor}')

    # Section 11 — births/deaths absolute counts (anchor at last existing
    # section-11 row "Младенческая смертность")
    if find_row_with_text(ws, LABEL_COL, 'Рождений (всего)'):
        print('  city: births/deaths counts already present, skipping')
    else:
        anchor11 = find_row_with_text(ws, LABEL_COL, 'Младенческая смертность')
        if not anchor11:
            anchor11 = find_row_with_text(ws, LABEL_COL, 'Медсестёр')
        if not anchor11:
            raise RuntimeError('Could not locate section 11 anchor')
        # Insert two new rows AFTER the natural-increase row
        ws.insert_rows(anchor11 + 1, amount=2)
        ws.cell(row=anchor11 + 1, column=LABEL_COL).value = 'Рождений (всего)'
        ws.cell(row=anchor11 + 1, column=UNIT_COL).value = 'чел.'
        ws.cell(row=anchor11 + 1, column=EX_COL).value = 6900
        copy_row_style(ws, anchor11, anchor11 + 1, ws.max_column)
        ws.cell(row=anchor11 + 2, column=LABEL_COL).value = 'Смертей (всего)'
        ws.cell(row=anchor11 + 2, column=UNIT_COL).value = 'чел.'
        ws.cell(row=anchor11 + 2, column=EX_COL).value = 1500
        copy_row_style(ws, anchor11, anchor11 + 2, ws.max_column)
        print(f'  city: inserted 2 vital-count rows after row {anchor11}')

    wb.save(path)
    print(f'  city: saved {path}')


# ────────────────────────────────────────────────────────────────────────
# 2. Region Excel — append new section "Темпы реального роста по секторам"
# ────────────────────────────────────────────────────────────────────────
def extend_region():
    path = os.path.join(HERE, 'golden_mart_region.xlsx')
    wb = o.load_workbook(path)
    ws = wb['Регионы']

    LABEL_COL = 2
    UNIT_COL  = 3
    EX_COL    = 4

    # Section title row probe
    if find_row_with_text(ws, LABEL_COL, '20. Темпы реального роста'):
        print('  region: real-growth section already present, skipping')
        wb.save(path)
        return

    # Find the last non-empty row to know where to append
    last = ws.max_row
    while last > 0 and all(
        ws.cell(row=last, column=c).value in (None, '') for c in range(1, ws.max_column + 1)
    ):
        last -= 1

    # Sample style row — pick a known data row to copy formatting from
    style_anchor = find_row_with_text(ws, LABEL_COL, 'ВРП')
    if not style_anchor:
        style_anchor = last  # fallback

    # Append blank row + section title + 6 sector rows
    sector_rows = [
        ('Промышленность – реальный рост',     '%', 107.3),
        ('Услуги – реальный рост',             '%', 108.6),
        ('Розничная торговля – реальный рост', '%', 111.1),
        ('Строительство – реальный рост',      '%', 117.4),
        ('Сельское хозяйство – реальный рост', '%', 105.4),
        ('ВРП – реальный рост',                '%', 108.1),
    ]

    base = last + 2  # blank row then section title
    ws.cell(row=base, column=LABEL_COL).value = '20. Темпы реального роста по секторам'
    copy_row_style(ws, style_anchor, base, ws.max_column)
    for i, (label, unit, example) in enumerate(sector_rows):
        r = base + 1 + i
        ws.cell(row=r, column=LABEL_COL).value = label
        ws.cell(row=r, column=UNIT_COL).value = unit
        ws.cell(row=r, column=EX_COL).value = example
        copy_row_style(ws, style_anchor, r, ws.max_column)

    wb.save(path)
    print(f'  region: appended section 20 at row {base}, saved {path}')


if __name__ == '__main__':
    print('Extending Golden Mart schemas:')
    extend_city()
    extend_region()
    print('Done.')
