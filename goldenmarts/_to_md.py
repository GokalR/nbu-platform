"""Convert each Golden Mart Excel into a structured MD spec.

Each Excel has 3 sheets: Инструкция, Глоссарий, and the data sheet. We
flatten the data sheet into Markdown with one section per "N. Section
title" row and one row per attribute (indicator | unit | example).

The MD doubles as (a) human-readable schema docs and (b) the canonical
list of attributes the unified dashboard / future admin form supports.
"""
from __future__ import annotations
import os
import openpyxl as o
import re

LEVELS = [
    ('country', 'golden_mart_country.xlsx', 'Узбекистан', 'Уровень страны (Узбекистан)'),
    ('region',  'golden_mart_region.xlsx',  'Регионы',   'Уровень области (региона)'),
    ('city',    'golden_mart_city.xlsx',    'Города и туманы', 'Уровень города/тумана'),
]

SECTION_RE = re.compile(r'^\d+\.\s+')


def cells(row):
    return [c.value if c.value is not None else '' for c in row]


def is_section(label: str) -> bool:
    return bool(SECTION_RE.match(str(label).strip()))


def convert(level_key: str, fn: str, sheet_name: str, title: str) -> str:
    wb = o.load_workbook(os.path.join(os.path.dirname(__file__), fn), data_only=True)
    ws = wb[sheet_name]
    glossary_ws = wb['Глоссарий']

    out = []
    out.append(f'# Golden Mart — {title}')
    out.append('')
    out.append(f'Источник: `{fn}`, лист `{sheet_name}`. Период: 2021–2025 + план 2026.')
    out.append('Каждая строка = атрибут, который заполняется значением за каждый год.')
    out.append('')
    out.append('## Содержание')
    out.append('')

    # First pass — collect sections
    sections = []
    cur = None
    for row in ws.iter_rows(values_only=False):
        vals = cells(row)
        # find first non-empty cell — that's the indicator label
        label = next((str(v).strip() for v in vals if v not in (None, '')), '')
        if not label:
            continue
        if is_section(label):
            cur = {'title': label, 'attrs': []}
            sections.append(cur)
            continue
        if cur is None:
            continue
        # data row: label | unit | example | 2021..2026
        non_empty = [str(v).strip() for v in vals if v not in (None, '')]
        if len(non_empty) < 2:
            continue
        # Skip header/intro rows
        if label in ('Показатель', 'Страна:', 'Область:', 'Город / туман:'):
            continue
        # Try to find unit + example by re-scanning row positions
        row_vals = vals
        # data sheets have leading blank columns; collapse them
        row_vals = [v for v in row_vals if v not in (None, '')]
        unit    = str(row_vals[1]).strip() if len(row_vals) > 1 else ''
        example = str(row_vals[2]).strip() if len(row_vals) > 2 else ''
        cur['attrs'].append({'label': label, 'unit': unit, 'example': example})

    for i, s in enumerate(sections, 1):
        anchor = re.sub(r'[^a-zа-я0-9]+', '-', s['title'].lower()).strip('-')
        out.append(f'{i}. [{s["title"]}](#{anchor})')
    out.append('')
    out.append('---')
    out.append('')

    for s in sections:
        out.append(f'## {s["title"]}')
        out.append('')
        out.append('| Показатель | Единица | Пример |')
        out.append('|---|---|---|')
        for a in s['attrs']:
            label   = a['label'].replace('|', '\\|')
            unit    = a['unit'].replace('|', '\\|')
            example = a['example'].replace('|', '\\|')
            out.append(f'| {label} | {unit} | {example} |')
        out.append('')

    # Glossary — short, only for country file (same content in all three)
    if level_key == 'country':
        out.append('---')
        out.append('')
        out.append('## Глоссарий (общий для всех уровней)')
        out.append('')
        out.append('| Термин | Определение |')
        out.append('|---|---|')
        for row in glossary_ws.iter_rows(values_only=True):
            vals = [v for v in row if v not in (None, '')]
            if len(vals) < 2:
                continue
            term, definition = str(vals[0]).strip(), str(vals[1]).strip()
            if term in ('Глоссарий', 'Термин'):
                continue
            term_e = term.replace('|', '\\|')
            def_e = definition.replace('|', '\\|')
            out.append(f'| {term_e} | {def_e} |')
        out.append('')

    return '\n'.join(out)


if __name__ == '__main__':
    here = os.path.dirname(__file__)
    for level_key, fn, sheet_name, title in LEVELS:
        md = convert(level_key, fn, sheet_name, title)
        out_path = os.path.join(here, f'GM_{level_key}.md')
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(md)
        print(f'  wrote {out_path}  ({len(md)} chars)')
